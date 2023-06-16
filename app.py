import pandas as pd
import warnings
from flask import Flask, render_template, request,redirect,url_for
import openpyxl
import random 
import smtplib
glovar =['0','0']
workbook = openpyxl.load_workbook('data\credentials.xlsx')
app = Flask(__name__)


@app.route('/', methods=['GET','POST'])
def login():
    if request.method=='POST':
    # Load the Excel sheet
          
        sheet = workbook.active
        username = request.form['username']
        password = request.form['password']

    # Check if the credentials match
        success = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(username) and str(row[1]) == str(password):
                username = str(row[2])
                success = True
                break

    # Redirect to the dashboard if the credentials are correct
        if success:
            return render_template('homepage.html', uname=username)
        else:
            return render_template('login.html', error_message="Invalid username or password")
    return render_template('login.html')

@app.route('/forgotpassword',methods=['GET','POST'])
def forgotPassword():
    if request.method=='POST':
        emailid = request.form['email']
        df = pd.read_excel('data\credentials.xlsx')
        if emailid not in df['username'].values:
             return render_template('forgot-password.html', error_message="Email not registered. Please check again.")
        elif emailid in df['username'].values:
            OTP=random.randint(100000, 999999)
            glovar[0] = OTP
            glovar[1]= emailid
            otp = str(OTP) + " is your OTP" #this store message and otp as strings  
            msg= otp 
            s = smtplib.SMTP('smtp.gmail.com', 587)
            s.starttls()
            s.login("kovi.pavan.kumar@gmail.com", "tlvurhteqfowztsj")
            user="kovi.pavan.kumar@gmail.com"
            s.sendmail(user,emailid,msg)
            s.quit()
            return redirect(url_for("update"))
    return render_template('forgot-password.html')

@app.route('/update',methods=['GET','POST'])
def update():
    if request.method=='POST':
        otp = request.form['otp']
        if str(glovar[0]) == str(otp):
            new_password = request.form['password']
            df = pd.read_excel('data\credentials.xlsx')

            df.loc[df['username'] == glovar[1], 'password'] = new_password

            df.to_excel('data\credentials.xlsx', index=False)
            return render_template('otp.html', msg="Password updated successfully")
        elif str(glovar[0])==otp:
            return render_template('otp.html', msg="OTP Mismatched")            

    return render_template('otp.html', msg="OTP sent succesfully")


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method=='POST':
         
        sheet = workbook.active
        existing_usernames = [cell.value for cell in sheet['A'][1:]]

        new_email = request.form['email']
        new_name = request.form['username']
        password = request.form['password']
        

        if new_email in existing_usernames:
            return render_template('register.html', existance="Email already exists")
        
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=new_email)
        sheet.cell(row=next_row, column=2, value=password)
        sheet.cell(row=next_row, column=3, value=new_name)
        workbook.save('data\credentials.xlsx')
        return render_template('register.html', existance="ACCOUNT SUCCESSFULLY CREATED")
    return render_template('register.html')

@app.route('/feedback', methods=['GET', 'POST'])
def feedback():
    if request.method =="POST":
        
        ml = openpyxl.load_workbook('data\ml.xlsx')
        next_row = ml.active.max_row + 1
        ml.active.cell(row=next_row, column=1, value=request.form['regno'])
        ml.active.cell(row=next_row, column=2, value=request.form['ml-com'])
        ml.active.cell(row=next_row, column=3, value=request.form['ml-know'])
        ml.active.cell(row=next_row, column=4, value=request.form['ml-Board'])
        ml.active.cell(row=next_row, column=5, value=request.form['ml-tech'])
        ml.active.cell(row=next_row, column=6, value=request.form['ml-pun'])
        ml.active.cell(row=next_row, column=7, value=request.form['ml-tm'])
        ml.active.cell(row=next_row, column=8, value=request.form['ml-pe'])
        ml.active.cell(row=next_row, column=9, value=request.form['ml-oa'])
        ml.save('data\ml.xlsx')

        cd = openpyxl.load_workbook('data\cd.xlsx')
        nex_row = cd.active.max_row + 1
        cd.active.cell(row=nex_row, column=1, value=request.form['regno'])
        cd.active.cell(row=nex_row, column=2, value=request.form['cd-com'])
        cd.active.cell(row=nex_row, column=3, value=request.form['cd-know'])
        cd.active.cell(row=nex_row, column=4, value=request.form['cd-Board'])
        cd.active.cell(row=nex_row, column=5, value=request.form['cd-tech'])
        cd.active.cell(row=nex_row, column=6, value=request.form['cd-pun'])
        cd.active.cell(row=nex_row, column=7, value=request.form['cd-tm'])
        cd.active.cell(row=nex_row, column=8, value=request.form['cd-pe'])
        cd.active.cell(row=nex_row, column=9, value=request.form['cd-oa'])
        cd.save('data\cd.xlsx')
      
        cns = openpyxl.load_workbook('data\cns.xlsx')
        ne_row = cns.active.max_row + 1
        cns.active.cell(row=ne_row, column=1, value=request.form['regno'])
        cns.active.cell(row=ne_row, column=2, value=request.form['cns-com'])
        cns.active.cell(row=ne_row, column=3, value=request.form['cns-know'])
        cns.active.cell(row=ne_row, column=4, value=request.form['cns-Board'])
        cns.active.cell(row=ne_row, column=5, value=request.form['cns-tech'])
        cns.active.cell(row=ne_row, column=6, value=request.form['cns-pun'])
        cns.active.cell(row=ne_row, column=7, value=request.form['cns-tm'])
        cns.active.cell(row=ne_row, column=8, value=request.form['cns-pe'])
        cns.active.cell(row=ne_row, column=9, value=request.form['cns-oa'])
        cns.save('data\cns.xlsx')

        iot = openpyxl.load_workbook('data\iot.xlsx')
        ne_row1 = iot.active.max_row + 1
        iot.active.cell(row=ne_row1, column=1, value=request.form['regno'])
        iot.active.cell(row=ne_row1, column=2, value=request.form['iot-com'])
        iot.active.cell(row=ne_row1, column=3, value=request.form['iot-know'])
        iot.active.cell(row=ne_row1, column=4, value=request.form['iot-Board'])
        iot.active.cell(row=ne_row1, column=5, value=request.form['iot-tech'])
        iot.active.cell(row=ne_row1, column=6, value=request.form['iot-pun'])
        iot.active.cell(row=ne_row1, column=7, value=request.form['iot-tm'])
        iot.active.cell(row=ne_row1, column=8, value=request.form['iot-pe'])
        iot.active.cell(row=ne_row1, column=9, value=request.form['iot-oa'])
        iot.save('data\iot.xlsx')       

        ooad = openpyxl.load_workbook('data\ooad.xlsx')
        ne_row2 = ooad.active.max_row + 1
        ooad.active.cell(row=ne_row2, column=1, value=request.form['regno'])
        ooad.active.cell(row=ne_row2, column=2, value=request.form['ooad-com'])
        ooad.active.cell(row=ne_row2, column=3, value=request.form['ooad-know'])
        ooad.active.cell(row=ne_row2, column=4, value=request.form['ooad-Board'])
        ooad.active.cell(row=ne_row2, column=5, value=request.form['ooad-tech'])
        ooad.active.cell(row=ne_row2, column=6, value=request.form['ooad-pun'])
        ooad.active.cell(row=ne_row2, column=7, value=request.form['ooad-tm'])
        ooad.active.cell(row=ne_row2, column=8, value=request.form['ooad-pe'])
        ooad.active.cell(row=ne_row2, column=9, value=request.form['ooad-oa'])
        ooad.save('data\ooad.xlsx')
        return render_template('success.html')

    
    return render_template('feedback.html')


if __name__ == '__main__':
    app.run()
